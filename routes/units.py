from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from models import db, Unit, Customer, LedgerEntry, Account, JournalEntry
from sqlalchemy import func
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from io import BytesIO

# Centralized Excel Styles (Copying from reports for now)
TITLE_FONT = Font(size=16, bold=True, color="1E293B")
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

units_bp = Blueprint('units', __name__)

@units_bp.route('/units')
def list_units():
    units = Unit.query.all()
    customers = Customer.query.all()
    return render_template('units.html', units=units, customers=customers)

@units_bp.route('/units/add', methods=['POST'])
def add_unit():
    unit_number = request.form.get('unit_number')
    new_unit = Unit(unit_number=unit_number, monthly_charge=0)
    db.session.add(new_unit)
    db.session.commit()
    flash('Unit added successfully', 'success')
    return redirect(url_for('units.list_units'))

@units_bp.route('/customers/add', methods=['POST'])
def add_customer():
    name = request.form.get('name')
    phone = request.form.get('phone')
    whatsapp_number = request.form.get('whatsapp_number')
    address = request.form.get('address')
    unit_id = request.form.get('unit_id')
    
    new_customer = Customer(name=name, phone=phone, whatsapp_number=whatsapp_number, address=address)
    db.session.add(new_customer)
    db.session.flush()
    
    if unit_id:
        unit = Unit.query.get(unit_id)
        if unit:
            unit.customer_id = new_customer.id
            unit.status = 'occupied'
            
    db.session.commit()
    flash('Customer added successfully', 'success')
    return redirect(url_for('units.list_units'))

@units_bp.route('/units/edit/<int:id>', methods=['POST'])
def edit_unit(id):
    unit = Unit.query.get_or_404(id)
    unit.unit_number = request.form.get('unit_number')
    db.session.commit()
    flash('Unit updated', 'success')
    return redirect(url_for('units.list_units'))

@units_bp.route('/units/delete/<int:id>', methods=['POST'])
def delete_unit(id):
    unit = Unit.query.get_or_404(id)
    db.session.delete(unit)
    db.session.commit()
    flash('Unit deleted', 'success')
    return redirect(url_for('units.list_units'))

@units_bp.route('/customers/edit/<int:id>', methods=['POST'])
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    customer.name = request.form.get('name')
    customer.phone = request.form.get('phone')
    customer.whatsapp_number = request.form.get('whatsapp_number')
    customer.address = request.form.get('address')
    db.session.commit()
    flash('Customer updated', 'success')
    return redirect(url_for('units.list_units'))

@units_bp.route('/customers/profile/<int:id>')
def customer_profile(id):
    customer = Customer.query.get_or_404(id)
    
    # Date filtering
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    today = datetime.now()
    default_from = today.replace(day=1).strftime('%Y-%m-%d')
    default_to = today.strftime('%Y-%m-%d')
    
    f_str = from_date_str if from_date_str else default_from
    t_str = to_date_str if to_date_str else default_to
    
    f_date = datetime.strptime(f_str, '%Y-%m-%d')
    t_date = datetime.strptime(t_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    
    # Calculate Arrears/Opening Balance (Everything before from_date)
    ar_acc = Account.query.filter(Account.code.like('39%')).first()
    opening_balance = 0
    if ar_acc:
        d = db.session.query(func.sum(LedgerEntry.debit)).join(JournalEntry).filter(
            LedgerEntry.customer_id == id,
            LedgerEntry.account_id == ar_acc.id,
            JournalEntry.date < f_date
        ).scalar() or 0
        c = db.session.query(func.sum(LedgerEntry.credit)).join(JournalEntry).filter(
            LedgerEntry.customer_id == id,
            LedgerEntry.account_id == ar_acc.id,
            JournalEntry.date < f_date
        ).scalar() or 0
        opening_balance = d - c

    # History within range
    history = LedgerEntry.query.join(JournalEntry).filter(
        LedgerEntry.customer_id == id,
        JournalEntry.date >= f_date,
        JournalEntry.date <= t_date
    ).order_by(JournalEntry.date.asc(), LedgerEntry.id.asc()).all()
    
    # Calculate current overall balance (total)
    balance = 0
    if ar_acc:
        debits = db.session.query(func.sum(LedgerEntry.debit)).filter_by(customer_id=id, account_id=ar_acc.id).scalar() or 0
        credits = db.session.query(func.sum(LedgerEntry.credit)).filter_by(customer_id=id, account_id=ar_acc.id).scalar() or 0
        balance = debits - credits
        
    return render_template('customer_profile.html', 
                           customer=customer, 
                           history=history, 
                           balance=balance, 
                           opening_balance=opening_balance,
                           from_date=f_str, 
                           to_date=t_str)

@units_bp.route('/customers/profile/<int:id>/export')
def export_customer_ledger(id):
    customer = Customer.query.get_or_404(id)
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    f_date = datetime.strptime(from_date_str, '%Y-%m-%d') if from_date_str else datetime.now().replace(day=1)
    t_date = datetime.strptime(to_date_str, '%Y-%m-%d') if to_date_str else datetime.now()
    t_date = t_date.replace(hour=23, minute=59, second=59)
    
    # Get ar account
    ar_acc = Account.query.filter(Account.code.like('39%')).first()
    
    # Opening Balance
    opening_balance = 0
    if ar_acc:
        d = db.session.query(func.sum(LedgerEntry.debit)).join(JournalEntry).filter(
            LedgerEntry.customer_id == id,
            LedgerEntry.account_id == ar_acc.id,
            JournalEntry.date < f_date
        ).scalar() or 0
        c = db.session.query(func.sum(LedgerEntry.credit)).join(JournalEntry).filter(
            LedgerEntry.customer_id == id,
            LedgerEntry.account_id == ar_acc.id,
            JournalEntry.date < f_date
        ).scalar() or 0
        opening_balance = d - c

    # Transactions
    history = LedgerEntry.query.join(JournalEntry).filter(
        LedgerEntry.customer_id == id,
        JournalEntry.date >= f_date,
        JournalEntry.date <= t_date
    ).order_by(JournalEntry.date.asc(), LedgerEntry.id.asc()).all()

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resident Ledger"
    
    # Header Information
    ws.merge_cells('A1:E1')
    ws['A1'] = "RESIDENT LEDGER STATEMENT"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Resident: {customer.name}"
    ws['A2'].font = Font(bold=True)
    ws['D2'] = f"Period: {f_date.strftime('%d %b %Y')} - {t_date.strftime('%d %b %Y')}"
    ws['D2'].font = Font(bold=True)
    
    unit = customer.units[0] if customer.units else None
    ws['A3'] = f"Unit: {unit.unit_number if unit else 'N/A'}"
    ws['D3'] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    
    # Table Header
    headers = ['Date', 'Description', 'Debit', 'Credit', 'Balance']
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = text
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
        
    # Opening Balance Row
    ws['A6'] = f_date.strftime('%d %b %Y')
    ws['B6'] = "OPENING BALANCE"
    ws['B6'].font = Font(italic=True)
    ws['E6'] = opening_balance
    ws['E6'].number_format = '#,##0.00'
    ws['E6'].font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=6, column=col).border = THIN_BORDER

    # Records
    curr_balance = opening_balance
    row = 7
    for entry in history:
        ws.cell(row=row, column=1).value = entry.parent.date.strftime('%d %b %Y')
        ws.cell(row=row, column=2).value = entry.parent.description
        ws.cell(row=row, column=3).value = entry.debit if entry.debit > 0 else 0
        ws.cell(row=row, column=4).value = entry.credit if entry.credit > 0 else 0
        
        # Balance calculation depends on if it's an AR account or anything else
        # For resident ledger, we primarily care about AR balance
        if ar_acc and entry.account_id == ar_acc.id:
            curr_balance += (entry.debit - entry.credit)
        
        ws.cell(row=row, column=5).value = curr_balance
        
        # Styling
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if col in [3, 4, 5]:
                cell.number_format = '#,##0.00'
        row += 1

    # Autosize
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Ledger_{customer.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
