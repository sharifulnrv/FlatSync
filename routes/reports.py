from flask import Blueprint, render_template, request, send_file, current_app, redirect, url_for
from utils.pdf_generator import render_to_pdf
from models import db, Account, LedgerEntry, Customer, Unit, JournalEntry
import traceback
from sqlalchemy import func
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border

# Centralized Excel Styles for Consistency
TITLE_FONT = Font(size=16, bold=True, color="1E293B")
ADDR_FONT = Font(size=10, color="64748B")
REPORT_TITLE_FONT = Font(size=12, bold=True, color="4F46E5")
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

reports_bp = Blueprint('reports', __name__)

def get_dates():
    from_date_str = request.args.get('from_date')
    to_date_str = request.args.get('to_date')
    
    today = datetime.now()
    default_from = today.replace(day=1).strftime('%Y-%m-%d')
    default_to = today.strftime('%Y-%m-%d')
    
    # Use default if not provided or empty string
    f_str = from_date_str if from_date_str else default_from
    t_str = to_date_str if to_date_str else default_to
    
    f_date = datetime.strptime(f_str, '%Y-%m-%d')
    t_date = datetime.strptime(t_str, '%Y-%m-%d')
    
    return f_date, t_date, f_str, t_str

def autosize_workbook(ws, min_width=15, skip_rows=4):
    """Safely autosize columns in an openpyxl worksheet."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        
        for cell in col:
            if cell.row <= skip_rows: continue
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max(max_length + 5, min_width)

@reports_bp.route('/reports/multi-unit-ledger')
def multi_unit_ledger():
    from models import Customer, Unit, Account, MonthlyBill
    from sqlalchemy import func
    
    # Get all residents
    residents = Customer.query.all()
    
    ledger_data = []
    for res in residents:
        if not res.units: continue
        
        unit_balances = []
        total_res_balance = 0
        
        for unit in res.units:
            unit_due = db.session.query(func.sum(MonthlyBill.amount + MonthlyBill.penalty_amount - MonthlyBill.paid_amount))\
                .filter(MonthlyBill.unit_id == unit.id, MonthlyBill.status != 'paid').scalar() or 0
            
            unit_balances.append({
                'unit_number': unit.unit_number,
                'monthly_charge': unit.monthly_charge,
                'balance': float(unit_due)
            })
            total_res_balance += float(unit_due)
            
        ledger_data.append({
            'resident_name': res.name,
            'phone': res.phone,
            'units': unit_balances,
            'total_balance': total_res_balance
        })
        
    return render_template('reports/multi_unit_ledger.html', ledger_data=ledger_data)

@reports_bp.route('/reports/aging')
def ar_aging_report():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    
    # NEW COA: 3930 is Service Charge Receivable
    ar_acc = Account.query.filter_by(code='3930').first()
    if not ar_acc:
        # Fallback or check for main AR range
        ar_acc = Account.query.filter(Account.code.like('39%'), Account.type == 'asset').first()
        
    if not ar_acc:
        return "A/R Account not found", 404
        
    # Base query for ledger entries in A/R
    q = db.session.query(LedgerEntry).filter(LedgerEntry.account_id == ar_acc.id)
    
    q = q.join(JournalEntry).filter(JournalEntry.date >= from_date)
    q = q.filter(JournalEntry.date <= to_date)
        
    all_entries = q.all()
    
    # Calculate balances per customer
    customer_balances = {}
    for entry in all_entries:
        cid = entry.customer_id
        if not cid: continue
        if cid not in customer_balances:
            customer_balances[cid] = {'customer': entry.customer or Customer.query.get(cid), 'balance': 0, 'oldest_date': None, 'transactions': []}
        
        customer_balances[cid]['balance'] += (entry.debit - entry.credit)
        customer_balances[cid]['transactions'].append(entry)
        
        entry_date = entry.parent.date
        if not customer_balances[cid]['oldest_date'] or entry_date < customer_balances[cid]['oldest_date']:
            customer_balances[cid]['oldest_date'] = entry_date

    aging_data = []
    now = datetime.now()
    
    for cid, data in customer_balances.items():
        if data['balance'] <= 0: continue
        
        days_old = (now - data['oldest_date']).days if data['oldest_date'] else 0
        category = "0-30 days"
        if days_old > 90: category = "90+ days"
        elif days_old > 60: category = "61-90 days"
        elif days_old > 30: category = "31-60 days"
        
        aging_data.append({
            'customer': data['customer'].name,
            'unit': data['customer'].units[0].unit_number if data['customer'].units else 'N/A',
            'balance': data['balance'],
            'days': days_old,
            'category': category,
            'transactions': data['transactions']
        })
        
    return render_template('aging_report.html', aging_data=aging_data, from_date=from_date_str, to_date=to_date_str)

@reports_bp.route('/reports/daily-cash')
def daily_cash_report():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    
    # NEW COA: 3x are Assets, 31xx are Cash/Bank
    # Group by date for the summary
    q_stats = db.session.query(
        func.date(JournalEntry.date).label('day'),
        func.sum(LedgerEntry.debit).label('total_in'),
        func.sum(LedgerEntry.credit).label('total_out')
    ).join(LedgerEntry, JournalEntry.id == LedgerEntry.journal_id)\
    .join(Account, LedgerEntry.account_id == Account.id)\
    .filter(Account.code.like('31%')) # 31xx are liquid cash/bank
    
    # Detailed query for individual transactions
    q_details = db.session.query(LedgerEntry).join(JournalEntry).join(Account, LedgerEntry.account_id == Account.id).filter(Account.code.like('31%'))

    # EXCLUDE Event cash from the main Association Daily Cash report
    q_stats = q_stats.filter(LedgerEntry.event_id == None)
    q_details = q_details.filter(LedgerEntry.event_id == None)

    q_stats = q_stats.filter(JournalEntry.date >= from_date)
    q_details = q_details.filter(JournalEntry.date >= from_date)
    q_stats = q_stats.filter(JournalEntry.date <= to_date)
    q_details = q_details.filter(JournalEntry.date <= to_date)
        
    daily_stats = q_stats.group_by(func.date(JournalEntry.date)).order_by(func.date(JournalEntry.date).desc()).all()
    transactions = q_details.order_by(JournalEntry.date.desc()).all()
    
    return render_template('daily_cash_report.html', stats=daily_stats, transactions=transactions, from_date=from_date_str, to_date=to_date_str)

@reports_bp.route('/reports/ledger')
def ledger_report():
    account_id = request.args.get('account_id', type=int)
    from_date, to_date, from_date_str, to_date_str = get_dates()
    
    accounts = Account.query.order_by(Account.code).all()
    target_account = Account.query.get(account_id) if account_id else None
    
    entries = []
    if target_account:
        q = LedgerEntry.query.join(JournalEntry).filter(LedgerEntry.account_id == target_account.id)
        
        # Ledger isolation
        event_id = request.args.get('event_id', type=int)
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)

        q = q.filter(JournalEntry.date >= from_date)
        q = q.filter(JournalEntry.date <= to_date)
        entries = q.order_by(JournalEntry.date.asc()).all()
        
    return render_template('ledger_report.html', 
                           accounts=accounts, 
                           target_account=target_account, 
                           entries=entries,
                           from_date=from_date_str, 
                           to_date=to_date_str)

@reports_bp.route('/reports')
def report_index():
    return render_template('reports_index.html')

@reports_bp.route('/reports/pnl')
def pnl_statement():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    event_id = request.args.get('event_id', type=int)
    
    revenue_accounts = Account.query.filter_by(type='revenue').order_by(Account.code).all()
    expense_accounts = Account.query.filter_by(type='expense').order_by(Account.code).all()
    
    from models import Event
    event = Event.query.get(event_id) if event_id else None
    
    pnl_data = {'revenue': [], 'expense': [], 'total_revenue': 0, 'total_expense': 0, 
                'from_date': from_date_str, 'to_date': to_date_str, 'event': event}
    
    for acc in revenue_accounts:
        if acc.is_summary: continue
        q = db.session.query(LedgerEntry).filter(LedgerEntry.account_id == acc.id)
        
        # Filter by event or association
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)
            
        q = q.join(JournalEntry).filter(JournalEntry.date >= from_date)
        q = q.filter(JournalEntry.date <= to_date)
            
        entries = q.all()
        balance = sum(e.credit - e.debit for e in entries)
        if balance != 0:
            pnl_data['revenue'].append({'name': acc.name, 'code': acc.code, 'balance': balance})
            pnl_data['total_revenue'] += balance
            
    for acc in expense_accounts:
        if acc.is_summary: continue
        q = db.session.query(LedgerEntry).filter(LedgerEntry.account_id == acc.id)
        
        # Filter by event or association
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)
            
        q = q.join(JournalEntry).filter(JournalEntry.date >= from_date)
        q = q.filter(JournalEntry.date <= to_date)
            
        entries = q.all()
        balance = sum(e.debit - e.credit for e in entries)
        if balance != 0:
            pnl_data['expense'].append({'name': acc.name, 'code': acc.code, 'balance': balance})
            pnl_data['total_expense'] += balance
            
    pnl_data['net_profit'] = pnl_data['total_revenue'] - pnl_data['total_expense']
    
    return render_template('pnl_report.html', data=pnl_data)

@reports_bp.route('/reports/due-report')
def due_report():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    # Service Charge Receivable (3930) or any 39x
    target_accs = Account.query.filter(Account.code.like('39%')).all()
    acc_ids = [a.id for a in target_accs]
    
    results = db.session.query(
        Customer.name,
        func.sum(LedgerEntry.debit) - func.sum(LedgerEntry.credit)
    ).join(LedgerEntry, Customer.id == LedgerEntry.customer_id)\
     .join(JournalEntry)\
     .filter(LedgerEntry.account_id.in_(acc_ids), LedgerEntry.event_id == None)\
     .filter(JournalEntry.date >= from_date)\
     .filter(JournalEntry.date <= to_date)\
     .group_by(Customer.id, Customer.name).all()
    return render_template('due_report.html', results=[r for r in results if r[1] != 0], from_date=from_date_str, to_date=to_date_str)

@reports_bp.route('/reports/trial-balance')
def trial_balance():
    from_date, to_date, from_date_str, to_date_str = get_dates()

    all_accounts = Account.query.order_by(Account.code).all()
    results = []
    total_debit = 0
    total_credit = 0
    
    for acc in all_accounts:
        # Include summary accounts if they have direct postings, but usually they shouldn't.
        # However, to avoid Trial Balance mismatch when users DO post to them, we MUST include them.
            
        # Main logic: If not specifically looking for an event, exclude event-linked entries from main Trial Balance
        event_id = request.args.get('event_id', type=int)
        q = db.session.query(func.sum(LedgerEntry.debit), func.sum(LedgerEntry.credit)).filter_by(account_id=acc.id)
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)

        q = q.join(JournalEntry).filter(JournalEntry.date >= from_date, JournalEntry.date <= to_date)
        
        debit, credit = q.first()
        debit, credit = (debit or 0), (credit or 0)
        
        if debit != 0 or credit != 0:
            results.append({'code': acc.code, 'name': acc.name, 'debit': debit, 'credit': credit, 'type': acc.type, 'is_summary': acc.is_summary})
            total_debit += debit
            total_credit += credit
            
    return render_template('reports/trial_balance.html', 
                           results=results, 
                           total_debit=total_debit, 
                           total_credit=total_credit,
                           from_date=from_date_str,
                           to_date=to_date_str)

@reports_bp.route('/reports/balance-sheet')
def balance_sheet():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    event_id = request.args.get('event_id', type=int)
    target_date = to_date

    assets = []
    liabilities = []
    equity = []
    
    total_assets = 0
    total_liabilities = 0
    total_equity = 0
    
    accounts = Account.query.filter(Account.type.in_(['asset', 'liability', 'equity'])).order_by(Account.code).all()
    
    for acc in accounts:
        if acc.is_summary: continue
        
        q = db.session.query(func.sum(LedgerEntry.debit), func.sum(LedgerEntry.credit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id, JournalEntry.date <= target_date)
        
        # Filter by event or association
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)
        
        debit, credit = q.first()
        debit = debit or 0
        credit = credit or 0
        
        balance = (debit - credit) if acc.type in ['asset', 'expense'] else (credit - debit)
        if balance != 0:
            item = {'code': acc.code, 'name': acc.name, 'balance': balance}
            if acc.type == 'asset':
                assets.append(item)
                total_assets += balance
            elif acc.type == 'liability':
                liabilities.append(item)
                total_liabilities += balance
            elif acc.type == 'equity':
                equity.append(item)
                total_equity += balance
                
    # Calculate Retained Earnings (Net Profit from beginning to target_date)
    # MUST mirror the event isolation logic
    rev_q = db.session.query(func.sum(LedgerEntry.credit - LedgerEntry.debit)).join(JournalEntry).join(Account).filter(Account.type == 'revenue', JournalEntry.date <= target_date)
    exp_q = db.session.query(func.sum(LedgerEntry.debit - LedgerEntry.credit)).join(JournalEntry).join(Account).filter(Account.type == 'expense', JournalEntry.date <= target_date)
    
    if event_id:
        rev_q = rev_q.filter(LedgerEntry.event_id == event_id)
        exp_q = exp_q.filter(LedgerEntry.event_id == event_id)
    else:
        rev_q = rev_q.filter(LedgerEntry.event_id == None)
        exp_q = exp_q.filter(LedgerEntry.event_id == None)

    retained_earnings = (rev_q.scalar() or 0) - (exp_q.scalar() or 0)
    
    if retained_earnings != 0:
        equity.append({'code': 'RE', 'name': 'Retained Earnings (Net Profit)', 'balance': retained_earnings})
        total_equity += retained_earnings

    return render_template('reports/balance_sheet.html', 
                            assets=assets, liabilities=liabilities, equity=equity,
                            total_assets=total_assets, total_liabilities=total_liabilities, total_equity=total_equity,
                            target_date=to_date_str, from_date=from_date_str, to_date=to_date_str)

@reports_bp.route('/reports/service-revenue')
def service_revenue_report():
    from models import MonthlyBill
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)
    
    query = MonthlyBill.query
    if month: query = query.filter_by(month=month)
    if year: query = query.filter_by(year=year)
    bills = query.order_by(MonthlyBill.year.desc(), MonthlyBill.month.desc()).all()
    
    stats = {
        'billed_service': sum(b.amount for b in bills),
        'billed_penalty': sum(b.penalty_amount for b in bills),
        'total_paid': sum(b.paid_amount for b in bills),
        'total_due': sum(b.balance_due for b in bills)
    }
    
    # Other Revenues from Journal (hall, gym, scrap, events)
    # COA Range 4000-4999 are Revenue
    rev_accounts = Account.query.filter(Account.type == 'revenue', Account.code != '4100').all() # 4100 is likely main service charge
    other_revs = []
    for acc in rev_accounts:
        amount = db.session.query(func.sum(LedgerEntry.credit - LedgerEntry.debit))\
            .filter(LedgerEntry.account_id == acc.id, LedgerEntry.event_id == None).scalar() or 0
        if amount != 0:
            other_revs.append({'name': acc.name, 'amount': amount})

    months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    return render_template('reports/service_revenue_report.html', 
                           bills=bills, stats=stats, other_revs=other_revs,
                           selected_month=month, selected_year=year, months=months)
@reports_bp.route('/reports/export/customers')
def export_customers():
    try:
        # Fetch units with residents
        units = Unit.query.all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Property Inventory"
        
        # Association styling
        # Styles now centralized at module level
        
        # Association Branding
        company_name = current_app.config.get('COMPANY_NAME', 'Association')
        company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
        ws.append([company_name])
        ws.append([company_address])
        ws.append(["PROPERTY INVENTORY REPORT"])
        ws.append([])
        
        # Branding Header (Standardized merge)
        last_col = "G"
        ws.merge_cells(f'A1:{last_col}1')
        ws.merge_cells(f'A2:{last_col}2')
        ws.merge_cells(f'A3:{last_col}3')
        
        for row_idx in [1, 2, 3]:
            cell = ws.cell(row=row_idx, column=1)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx == 1: cell.font = TITLE_FONT
            elif row_idx == 2: cell.font = ADDR_FONT
            else: cell.font = REPORT_TITLE_FONT
    except Exception as e:
        import traceback
        with open("error.log", "a") as f:
            f.write(f"\n--- CUSTOMER EXPORT ERROR AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e

    # Modern Header Row (Row 5)
    headers = ["Unit Number", "Status", "Resident Name", "Phone", "Mailing Address"]
    ws.append(headers)
    
    header_fill = HEADER_FILL
    header_font = HEADER_FONT
    thin_border = THIN_BORDER
    
    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for unit in units:
        res_name = unit.resident.name if unit.resident else "N/A"
        res_phone = unit.resident.phone if unit.resident else "N/A"
        res_addr = unit.resident.address if unit.resident else "N/A"
        ws.append([unit.unit_number, unit.status.title(), res_name, res_phone, res_addr])

    autosize_workbook(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Property_Inventory_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@reports_bp.route('/reports/export/pnl')
def export_pnl():
    try:
        from_date, to_date, from_date_str, to_date_str = get_dates()
        event_id = request.args.get('event_id', type=int)
        
        revenue_accounts = Account.query.filter_by(type='revenue').order_by(Account.code).all()
        expense_accounts = Account.query.filter_by(type='expense').order_by(Account.code).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profit & Loss"
        
        # Style definitions inlined or centralized
        
        company_name = current_app.config.get('COMPANY_NAME', 'Association')
        company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
        ws.append([company_name])
        ws.append([company_address])
        ws.append(["PROFIT & LOSS STATEMENT"])
        ws.append([f"Reporting Period: {from_date_str} to {to_date_str}"])
        ws.append([]) # Spacer
        
        # Styling
        # Branding Header
        last_col = "G"
        for row_idx in [1, 2, 3, 4]:
            ws.merge_cells(f'A{row_idx}:{last_col}{row_idx}')
            cell = ws.cell(row=row_idx, column=1)
            cell.alignment = Alignment(horizontal="center")
            if row_idx == 1: cell.font = TITLE_FONT
            elif row_idx == 2: cell.font = ADDR_FONT
            else: cell.font = REPORT_TITLE_FONT
    except Exception as e:
        with open("error.log", "a") as f:
            f.write(f"\n--- PNL EXPORT ERROR AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e
    
    ws.append(["REVENUE"])
    total_rev = 0
    for acc in revenue_accounts:
        if acc.is_summary: continue
        q = db.session.query(LedgerEntry).filter(LedgerEntry.account_id == acc.id)
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)
        q = q.join(JournalEntry).filter(JournalEntry.date >= from_date)
        q = q.filter(JournalEntry.date <= to_date)
        balance = sum(e.credit - e.debit for e in q.all())
        if balance != 0:
            ws.append([f"{acc.code} - {acc.name}", balance])
            total_rev += balance
    ws.append(["Total Revenue", total_rev])
    ws.append([])

    ws.append(["EXPENSES"])
    total_exp = 0
    for acc in expense_accounts:
        if acc.is_summary: continue
        q = db.session.query(LedgerEntry).filter(LedgerEntry.account_id == acc.id)
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)
        q = q.join(JournalEntry).filter(JournalEntry.date >= from_date, JournalEntry.date <= to_date)
        balance = sum(e.debit - e.credit for e in q.all())
        if balance != 0:
            ws.append([f"{acc.code} - {acc.name}", balance])
            total_exp += balance
    ws.append(["Total Expenses", total_exp])
    ws.append([])
    ws.append(["NET PROFIT", total_rev - total_exp])

    autosize_workbook(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"PNL_Statement_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

@reports_bp.route('/reports/export/trial-balance')
def export_trial_balance():
    try:
        from_date, to_date, from_date_str, to_date_str = get_dates()
        all_accounts = Account.query.order_by(Account.code).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trial Balance"
        # Style definitions now centralized at module level
        company_name = current_app.config.get('COMPANY_NAME', 'Association')
        company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
        ws.append([company_name])
        ws.append([company_address])
        ws.append(["TRIAL BALANCE REPORT"])
        ws.append([f"As Of Period: {from_date_str} to {to_date_str}"])
        ws.append([])
        
        # Styling
        # Branding Header
        last_col = "G"
        for row_idx in [1, 2, 3, 4]:
            ws.merge_cells(f'A{row_idx}:{last_col}{row_idx}')
            cell = ws.cell(row=row_idx, column=1)
            cell.alignment = Alignment(horizontal="center")
            if row_idx == 1: cell.font = TITLE_FONT
            elif row_idx == 2: cell.font = ADDR_FONT
            else: cell.font = REPORT_TITLE_FONT
    except Exception as e:
        with open("error.log", "a") as f:
            f.write(f"\n--- TRIAL BALANCE EXPORT ERROR AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e
    
    ws.append(["Code", "Account Name", "Debit", "Credit"])
    
    t_debit = 0
    t_credit = 0
    for acc in all_accounts:
        if acc.is_summary: continue
        q = db.session.query(func.sum(LedgerEntry.debit), func.sum(LedgerEntry.credit)).filter_by(account_id=acc.id)
        event_id = request.args.get('event_id', type=int)
        if event_id:
            q = q.filter(LedgerEntry.event_id == event_id)
        else:
            q = q.filter(LedgerEntry.event_id == None)
        q = q.join(JournalEntry).filter(JournalEntry.date >= from_date)
        q = q.filter(JournalEntry.date <= to_date)
        d, c = q.first()
        d, c = (d or 0), (c or 0)
        if d != 0 or c != 0:
            ws.append([acc.code, acc.name, d, c])
            t_debit += d
            t_credit += c
    ws.append(["Total", "TOTAL", t_debit, t_credit])

    autosize_workbook(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"Trial_Balance_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

@reports_bp.route('/reports/export/balance-sheet')
def export_balance_sheet():
    try:
        from_date, to_date, from_date_str, to_date_str = get_dates()
        event_id = request.args.get('event_id', type=int)
        target_date = to_date # BS is usually as of date, we'll use 'to_date'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        # Style definitions now centralized at module level
        company_name = current_app.config.get('COMPANY_NAME', 'Association')
        company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
        ws.append([company_name])
        ws.append([company_address])
        ws.append(["BALANCE SHEET STATEMENT"])
        ws.append([f"As Of Date: {to_date_str}"])
        ws.append([])
        
        # Styling
        # Branding Header
        last_col = "G"
        for row_idx in [1, 2, 3, 4]:
            ws.merge_cells(f'A{row_idx}:{last_col}{row_idx}')
            cell = ws.cell(row=row_idx, column=1)
            cell.alignment = Alignment(horizontal="center")
            if row_idx == 1: cell.font = TITLE_FONT
            elif row_idx == 2: cell.font = ADDR_FONT
            else: cell.font = REPORT_TITLE_FONT
    except Exception as e:
        with open("error.log", "a") as f:
            f.write(f"\n--- BALANCE SHEET EXPORT ERROR AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e
    
    def add_section(title, acc_type, is_credit_balance):
        ws.append([title])
        total = 0
        accounts = Account.query.filter_by(type=acc_type).order_by(Account.code).all()
        for acc in accounts:
            if acc.is_summary: continue
            q = db.session.query(func.sum(LedgerEntry.debit), func.sum(LedgerEntry.credit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id, JournalEntry.date <= target_date)
            if event_id:
                q = q.filter(LedgerEntry.event_id == event_id)
            else:
                q = q.filter(LedgerEntry.event_id == None)
            d, c = q.first()
            d, c = (d or 0), (c or 0)
            bal = (c - d) if is_credit_balance else (d - c)
            if bal != 0:
                ws.append([acc.name, bal])
                total += bal
        return total

    # 1. Assets
    t_assets = add_section("ASSETS", "asset", False)
    ws.append(["TOTAL ASSETS", t_assets])
    ws.append([])
    
    # 2. Liabilities
    t_liabs = add_section("LIABILITIES", "liability", True)
    ws.append(["TOTAL LIABILITIES", t_liabs])
    ws.append([])
    
    # 3. Retained Earnings (Profit/Loss)
    rev_q = db.session.query(func.sum(LedgerEntry.credit - LedgerEntry.debit)).join(JournalEntry).join(Account).filter(Account.type == 'revenue', JournalEntry.date <= target_date)
    exp_q = db.session.query(func.sum(LedgerEntry.debit - LedgerEntry.credit)).join(JournalEntry).join(Account).filter(Account.type == 'expense', JournalEntry.date <= target_date)
    
    if event_id:
        rev_q = rev_q.filter(LedgerEntry.event_id == event_id)
        exp_q = exp_q.filter(LedgerEntry.event_id == event_id)
    else:
        rev_q = rev_q.filter(LedgerEntry.event_id == None)
        exp_q = exp_q.filter(LedgerEntry.event_id == None)
        
    net_earnings = (rev_q.scalar() or 0) - (exp_q.scalar() or 0)
    
    # 4. Equity
    ws.append(["EQUITY"])
    t_equity = add_section("Post-Direct Equity", "equity", True) # Usually empty unless direct equity entries exist
    ws.append(["Retained Earnings (P&L)", net_earnings])
    
    total_equity = t_equity + net_earnings
    ws.append(["TOTAL EQUITY", total_equity])
    ws.append([])
    
    ws.append(["TOTAL LIABILITIES & EQUITY", t_liabs + total_equity])

    autosize_workbook(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"Balance_Sheet_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

@reports_bp.route('/reports/export/ledger')
def export_ledger():
    account_id = request.args.get('account_id', type=int)
    from_date, to_date, from_date_str, to_date_str = get_dates()
    
    target_account = Account.query.get_or_404(account_id)
    
    q = LedgerEntry.query.join(JournalEntry).filter(LedgerEntry.account_id == target_account.id)
    q = q.filter(JournalEntry.date >= from_date)
    q = q.filter(JournalEntry.date <= to_date)
    entries = q.order_by(JournalEntry.date.asc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ledger_{target_account.code}"
    
    # Style definitions now centralized at module level
    company_name = current_app.config.get('COMPANY_NAME', 'Association')
    company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
    ws.append([company_name])
    ws.append([company_address])
    ws.append([f"DETAILED LEDGER REPORT: {target_account.code}"])
    ws.append([f"{target_account.name} | Period: {from_date_str} to {to_date_str}"])
    ws.append([])
    
    # Branding Header
    last_col = "G"
    for row_idx in [1, 2, 3, 4]:
        ws.merge_cells(f'A{row_idx}:{last_col}{row_idx}')
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = Alignment(horizontal="center")
        if row_idx == 1: cell.font = TITLE_FONT
        elif row_idx == 2: cell.font = ADDR_FONT
        else: cell.font = REPORT_TITLE_FONT
    
    ws.append(["Date", "Reference", "Description", "Debit", "Credit", "Balance"])
    
    balance = 0
    for entry in entries:
        if target_account.type in ['asset', 'expense']:
            balance += (entry.debit - entry.credit)
        else:
            balance += (entry.credit - entry.debit)
            
        ws.append([
            entry.parent.date.strftime('%Y-%m-%d'),
            entry.parent.reference or 'JR',
            entry.parent.description,
            entry.debit,
            entry.credit,
            balance
        ])

    try:
        autosize_workbook(ws, skip_rows=4)
        # The ledger can be long, auto-sizing is essential
    
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"Ledger_{target_account.code}_{datetime.now().strftime('%Y-%m-%d')}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        with open("error.log", "a") as f:
            f.write(f"\n--- LEDGER EXPORT ERROR (SAVE) AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e

@reports_bp.route('/reports/export/aging')
def export_aging():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    
    residents = Customer.query.all()
    aging_data = []
    
    ar_acc = Account.query.filter_by(code='3930').first()
    ar_id = ar_acc.id if ar_acc else None
    
    for res in residents:
        if not ar_id: continue
        q = LedgerEntry.query.join(JournalEntry).filter(LedgerEntry.account_id == ar_id, LedgerEntry.customer_id == res.id)
        # Using customer_id is much more reliable than description.contains(res.name)
        q = q.filter(JournalEntry.date >= from_date)
        q = q.filter(JournalEntry.date <= to_date)
        
        balance = sum(e.debit - e.credit for e in q.all())
        if balance > 0:
            first_unpaid = q.filter(LedgerEntry.debit > LedgerEntry.credit).order_by(JournalEntry.date.asc()).first()
            days = (datetime.utcnow() - first_unpaid.parent.date).days if first_unpaid else 0
            cat = "0-30 days" if days <= 30 else "31-60 days" if days <= 60 else "61-90 days" if days <= 90 else "90+ days"
            aging_data.append({'customer': res.name, 'balance': balance, 'days': days, 'category': cat})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR Aging"
    # Style definitions now centralized at module level
    company_name = current_app.config.get('COMPANY_NAME', 'Association')
    company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
    ws.append([company_name])
    ws.append([company_address])
    ws.append(["ACCOUNTS RECEIVABLE AGING REPORT"])
    ws.append([f"Aging Basis: {from_date_str} to {to_date_str}"])
    ws.append([])
    
    # Branding Header
    last_col = "G"
    for row_idx in [1, 2, 3, 4]:
        ws.merge_cells(f'A{row_idx}:{last_col}{row_idx}')
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = Alignment(horizontal="center")
        if row_idx == 1: cell.font = TITLE_FONT
        elif row_idx == 2: cell.font = ADDR_FONT
        else: cell.font = REPORT_TITLE_FONT
    
    ws.append(["Customer", "Days Overdue", "Category", "Balance"])
    
    for item in aging_data:
        ws.append([item['customer'], item['days'], item['category'], item['balance']])
        
    try:
        autosize_workbook(ws)
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"AR_Aging_{datetime.now().strftime('%Y-%m-%d')}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        with open("error.log", "a") as f:
            f.write(f"\n--- AGING EXPORT ERROR (SAVE) AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e

@reports_bp.route('/reports/export/daily-cash')
def export_daily_cash():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    
    q = LedgerEntry.query.join(Account).filter(Account.code.like('31%'), LedgerEntry.event_id == None)
    q = q.join(JournalEntry).filter(JournalEntry.date >= from_date)
    q = q.filter(JournalEntry.date <= to_date)
    entries = q.order_by(JournalEntry.date.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Cash"
    # Style definitions now centralized at module level
    company_name = current_app.config.get('COMPANY_NAME', 'Association')
    company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
    ws.append([company_name])
    ws.append([company_address])
    ws.append(["DAILY CASH FLOW REPORT"])
    ws.append([f"Audit Period: {from_date_str} to {to_date_str}"])
    ws.append([])
    
    # Branding Header
    last_col = "G"
    for row_idx in [1, 2, 3, 4]:
        ws.merge_cells(f'A{row_idx}:{last_col}{row_idx}')
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = Alignment(horizontal="center")
        if row_idx == 1: cell.font = TITLE_FONT
        elif row_idx == 2: cell.font = ADDR_FONT
        else: cell.font = REPORT_TITLE_FONT
    
    ws.append(["Date", "Account", "Description", "Inflow (+)", "Outflow (-)", "Net"])
    
    for e in entries:
        inflow = e.debit if e.debit > 0 else 0
        outflow = e.credit if e.credit > 0 else 0
        ws.append([e.parent.date.strftime('%Y-%m-%d'), e.account.name, e.parent.description, inflow, outflow, inflow - outflow])
        
    try:
        autosize_workbook(ws)
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"Daily_Cash_{datetime.now().strftime('%Y-%m-%d')}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        with open("error.log", "a") as f:
            f.write(f"\n--- DAILY CASH EXPORT ERROR (SAVE) AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e

@reports_bp.route('/reports/export/due-report')
def export_due_report():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    target_accs = Account.query.filter(Account.code.like('39%')).all()
    acc_ids = [a.id for a in target_accs]
    
    results = db.session.query(
        Customer.name,
        func.sum(LedgerEntry.debit) - func.sum(LedgerEntry.credit)
    ).join(LedgerEntry, Customer.id == LedgerEntry.customer_id)\
     .join(JournalEntry)\
     .filter(LedgerEntry.account_id.in_(acc_ids), LedgerEntry.event_id == None)\
     .filter(JournalEntry.date >= from_date)\
     .filter(JournalEntry.date <= to_date)\
     .group_by(Customer.id, Customer.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Due Report"
    # Style definitions now centralized at module level
    company_name = current_app.config.get('COMPANY_NAME', 'Association')
    company_address = current_app.config.get('COMPANY_ADDRESS', 'Address')
    ws.append([company_name])
    ws.append([company_address])
    ws.append(["CUSTOMER OUTSTANDING BALANCES"])
    ws.append([f"Reference Period: {from_date_str} to {to_date_str}"])
    ws.append([])
    
    # Branding Header
    last_col = "G"
    for row_idx in [1, 2, 3, 4]:
        ws.merge_cells(f'A{row_idx}:{last_col}{row_idx}')
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = Alignment(horizontal="center")
        if row_idx == 1: cell.font = TITLE_FONT
        elif row_idx == 2: cell.font = ADDR_FONT
        else: cell.font = REPORT_TITLE_FONT
    
    ws.append(["Customer Name", "Outstanding Amount"])
    
    try:
        for name, balance in results:
            if balance != 0:
                ws.append([name, balance])
                
        autosize_workbook(ws)
                
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name=f"Due_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        with open("error.log", "a") as f:
            f.write(f"\n--- DUE REPORT EXPORT ERROR (SAVE) AT {datetime.now()} ---\n")
            traceback.print_exc(file=f)
        raise e
@reports_bp.route('/reports/breakdown')
def account_breakdown():
    from_date, to_date, from_date_str, to_date_str = get_dates()
    acc_code = request.args.get('code')
    
    if not acc_code:
        return redirect(url_for('reports.pnl_statement'))
        
    account = Account.query.filter_by(code=acc_code).first_or_404()
    
    # Base query for ledger entries
    q = db.session.query(LedgerEntry).join(JournalEntry).filter(LedgerEntry.account_id == account.id)
    q = q.filter(JournalEntry.date >= from_date, JournalEntry.date <= to_date)
    
    # Exclude event transactions from the main breakdown
    q = q.filter(LedgerEntry.event_id == None)
    
    entries = q.all()
    
    # Breakdown by Resident (for income)
    resident_breakdown = {}
    # Breakdown by Party (for expenses)
    party_breakdown = {}
    
    for entry in entries:
        if entry.customer_id:
            res = entry.customer
            cid = entry.customer_id
            if cid not in resident_breakdown:
                resident_breakdown[cid] = {'name': res.name, 'unit': res.units[0].unit_number if res.units else 'N/A', 'total': 0}
            amount = (entry.credit - entry.debit) if account.type == 'revenue' else (entry.debit - entry.credit)
            resident_breakdown[cid]['total'] += amount
            
        if entry.party_id:
            pty = entry.party
            pid = entry.party_id
            if pid not in party_breakdown:
                party_breakdown[pid] = {'name': pty.name, 'type': pty.type, 'total': 0}
            amount = (entry.debit - entry.credit) if account.type == 'expense' else (entry.credit - entry.debit)
            party_breakdown[pid]['total'] += amount

    total_balance = sum((e.debit - e.credit) if account.type in ['asset', 'expense'] else (e.credit - e.debit) for e in entries)

    return render_template('reports/account_breakdown.html', 
                            account=account, 
                            entries=entries, 
                            resident_breakdown=resident_breakdown.values(),
                            party_breakdown=party_breakdown.values(),
                            total_balance=total_balance,
                            from_date=from_date_str, 
                            to_date=to_date_str)

@reports_bp.route('/reports/trial-balance/pdf')
def trial_balance_pdf():
    from models import Account, LedgerEntry
    f_date, t_date, f_str, t_str = get_dates()
    event_id = request.args.get('event_id', type=int)
    
    accounts = Account.query.filter(Account.is_summary == False).order_by(Account.code).all()
    results = []
    total_debit = 0
    total_credit = 0
    
    for acc in accounts:
        query = db.session.query(func.sum(LedgerEntry.debit), func.sum(LedgerEntry.credit))\
                          .join(JournalEntry)\
                          .filter(LedgerEntry.account_id == acc.id)
        if f_date: query = query.filter(JournalEntry.date >= f_date)
        if t_date: query = query.filter(JournalEntry.date <= t_date)
        if event_id: query = query.filter(LedgerEntry.event_id == event_id)
        else: query = query.filter(LedgerEntry.event_id == None)
        
        dr, cr = query.first()
        dr = dr or 0
        cr = cr or 0
        if dr != 0 or cr != 0:
            results.append({'code': acc.code, 'name': acc.name, 'debit': dr, 'credit': cr})
            total_debit += dr
            total_credit += cr

    from io import BytesIO
    pdf_content = render_to_pdf('reports/trial_balance.html', {
        'results': results, 'total_debit': total_debit, 'total_credit': total_credit,
        'from_date': f_str, 'to_date': t_str
    })
    if pdf_content:
        return send_file(BytesIO(pdf_content), download_name=f"Trial_Balance_{datetime.now().strftime('%Y%m%d')}.pdf", as_attachment=True)
    return "Error", 500

@reports_bp.route('/reports/pnl/pdf')
def pnl_statement_pdf():
    from models import Account, LedgerEntry, Event
    f_date, t_date, f_str, t_str = get_dates()
    event_id = request.args.get('event_id', type=int)
    event = Event.query.get(event_id) if event_id else None
    
    rev_accs = Account.query.filter(Account.type == 'revenue', Account.is_summary == False).all()
    exp_accs = Account.query.filter(Account.type == 'expense', Account.is_summary == False).all()
    
    revenue_items = []
    expense_items = []
    total_revenue = 0
    total_expense = 0
    
    for acc in rev_accs:
        bal = db.session.query(func.sum(LedgerEntry.credit) - func.sum(LedgerEntry.debit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id)
        if f_date: bal = bal.filter(JournalEntry.date >= f_date)
        if t_date: bal = bal.filter(JournalEntry.date <= t_date)
        if event_id: bal = bal.filter(LedgerEntry.event_id == event_id)
        else: bal = bal.filter(LedgerEntry.event_id == None)
        val = bal.scalar() or 0
        if val != 0:
            revenue_items.append({'code': acc.code, 'name': acc.name, 'balance': val})
            total_revenue += val
            
    for acc in exp_accs:
        bal = db.session.query(func.sum(LedgerEntry.debit) - func.sum(LedgerEntry.credit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id)
        if f_date: bal = bal.filter(JournalEntry.date >= f_date)
        if t_date: bal = bal.filter(JournalEntry.date <= t_date)
        if event_id: bal = bal.filter(LedgerEntry.event_id == event_id)
        else: bal = bal.filter(LedgerEntry.event_id == None)
        val = bal.scalar() or 0
        if val != 0:
            expense_items.append({'code': acc.code, 'name': acc.name, 'balance': val})
            total_expense += val

    from io import BytesIO
    pdf_content = render_to_pdf('pnl_report.html', {
        'data': {
            'revenue': revenue_items,
            'expense': expense_items,
            'total_revenue': total_revenue,
            'total_expense': total_expense,
            'net_profit': total_revenue - total_expense,
            'from_date': f_str,
            'to_date': t_str,
            'event': event
        }
    })
    if pdf_content:
        return send_file(BytesIO(pdf_content), download_name=f"PnL_Statement_{datetime.now().strftime('%Y%m%d')}.pdf", as_attachment=True)
    return "Error", 500

@reports_bp.route('/reports/multi-unit/pdf')
def multi_unit_ledger_pdf():
    from models import Customer
    from io import BytesIO
    residents = Customer.query.all()
    ledger_data = []
    for res in residents:
        units_data = []
        total_balance = 0
        for unit in res.units:
            bal = db.session.query(func.sum(LedgerEntry.debit) - func.sum(LedgerEntry.credit))\
                            .filter(LedgerEntry.customer_id == res.id, LedgerEntry.account_id == 14)\
                            .scalar() or 0
            units_data.append({'unit_number': unit.unit_number, 'balance': float(bal)})
            total_balance += float(bal)
        ledger_data.append({'resident_name': res.name, 'phone': res.phone, 'units': units_data, 'total_balance': total_balance})

    pdf_content = render_to_pdf('reports/multi_unit_ledger.html', {'ledger_data': ledger_data})
    if pdf_content:
        return send_file(BytesIO(pdf_content), download_name=f"Multi_Unit_Ledger_{datetime.now().strftime('%Y%m%d')}.pdf", as_attachment=True)
    return "Error", 500

@reports_bp.route('/reports/service-revenue/pdf')
def service_revenue_report_pdf():
    from io import BytesIO
    f_date, t_date, f_str, t_str = get_dates()
    rev_accs = Account.query.filter(Account.code.like('41%'), Account.is_summary == False).all()
    report_data = []
    for acc in rev_accs:
        billed = db.session.query(func.sum(LedgerEntry.credit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id)
        paid = db.session.query(func.sum(LedgerEntry.debit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id)
        if f_date: 
            billed = billed.filter(JournalEntry.date >= f_date)
            paid = paid.filter(JournalEntry.date >= f_date)
        if t_date: 
            billed = billed.filter(JournalEntry.date <= t_date)
            paid = paid.filter(JournalEntry.date <= t_date)
        report_data.append({'account_name': acc.name, 'billed': billed.scalar() or 0, 'paid': paid.scalar() or 0})

    pdf_content = render_to_pdf('reports/service_revenue_report.html', {
        'data': report_data, 'from_date': f_str, 'to_date': t_str
    })
    if pdf_content:
        return send_file(BytesIO(pdf_content), download_name=f"Service_Revenue_{datetime.now().strftime('%Y%m%d')}.pdf", as_attachment=True)
    return "Error", 500

@reports_bp.route('/reports/breakdown/pdf')
def account_breakdown_pdf():
    from io import BytesIO
    acc_id = request.args.get('account_id', type=int)
    from_date, to_date, f_str, t_str = get_dates()
    account = Account.query.get_or_404(acc_id)
    entries = LedgerEntry.query.join(JournalEntry).filter(LedgerEntry.account_id == acc_id).order_by(JournalEntry.date.desc())
    if from_date: entries = entries.filter(JournalEntry.date >= from_date)
    if to_date: entries = entries.filter(JournalEntry.date <= to_date)
    entries = entries.all()
    
    total_balance = sum((e.debit - e.credit) if account.type in ['asset', 'expense'] else (e.credit - e.debit) for e in entries)
    
    pdf_content = render_to_pdf('reports/account_breakdown.html', {
        'account': account, 'entries': entries, 'total_balance': total_balance,
        'from_date': f_str, 'to_date': t_str
    })
    if pdf_content:
        return send_file(BytesIO(pdf_content), download_name=f"Breakdown_{account.name}_{datetime.now().strftime('%Y%m%d')}.pdf", as_attachment=True)
    return "Error", 500

@reports_bp.route('/reports/balance-sheet/pdf')
def balance_sheet_pdf():
    from io import BytesIO
    target_date_str = request.args.get('to_date')
    if target_date_str:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    else:
        target_date = datetime.now().date()
    
    asset_accs = Account.query.filter(Account.type == 'asset', Account.is_summary == False).all()
    liab_accs = Account.query.filter(Account.type == 'liability', Account.is_summary == False).all()
    equity_accs = Account.query.filter(Account.type == 'equity', Account.is_summary == False).all()
    
    def get_bal(acc, date):
        dr = db.session.query(func.sum(LedgerEntry.debit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id, JournalEntry.date <= date).scalar() or 0
        cr = db.session.query(func.sum(LedgerEntry.credit)).join(JournalEntry).filter(LedgerEntry.account_id == acc.id, JournalEntry.date <= date).scalar() or 0
        return dr - cr if acc.type in ['asset', 'expense'] else cr - dr

    assets = [{'code': a.code, 'name': a.name, 'balance': get_bal(a, target_date)} for a in asset_accs]
    liabilities = [{'code': a.code, 'name': a.name, 'balance': get_bal(a, target_date)} for a in liab_accs]
    equity = [{'code': a.code, 'name': a.name, 'balance': get_bal(a, target_date)} for a in equity_accs]
    
    total_assets = sum(a['balance'] for a in assets)
    total_liabilities = sum(l['balance'] for l in liabilities)
    total_equity = sum(e['balance'] for e in equity)

    pdf_content = render_to_pdf('reports/balance_sheet.html', {
        'assets': [a for a in assets if a['balance'] != 0],
        'liabilities': [l for l in liabilities if l['balance'] != 0],
        'equity': [e for e in equity if e['balance'] != 0],
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'target_date': target_date.strftime('%B %d, %Y'),
        'to_date': target_date.strftime('%Y-%m-%d')
    })
    if pdf_content:
        return send_file(BytesIO(pdf_content), download_name=f"Balance_Sheet_{target_date.strftime('%Y%m%d')}.pdf", as_attachment=True)
    return "Error", 500
